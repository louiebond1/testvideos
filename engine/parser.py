"""Reads an EX3 RCM test workbook (Excel) and returns a list of TestScenario objects."""

import os
from typing import List, Optional
import openpyxl

from models.dataclasses import TestScenario, TestStep

# Column indices (0-based)
COL_ROW_NUM = 0       # A — row number within scenario
COL_SCRIPT_ID = 1     # B — scenario ID e.g. LOGIN-100
COL_SCENARIO = 2      # C — scenario name (only on first row of each group)
COL_ROLE = 3          # D — role e.g. Recruiter
COL_STEP = 4          # E — step number
COL_ACTION = 5        # F — free-text action description
COL_TEST_DATA = 6     # G — input / test data
COL_EXPECTED = 7      # H — expected result
COL_STATUS = 8        # I — status (To Be Tested / Pass / Fail)
COL_COMMENTS = 9      # J — comments


def _is_section_header(row: tuple) -> bool:
    """Return True if this row is a section divider (►/◀ marker), not a test step."""
    first = row[COL_ROW_NUM]
    if not isinstance(first, str):
        return False
    stripped = first.strip()
    return stripped.startswith("►") or stripped.startswith("◄") or stripped.startswith("◀") or stripped.startswith("PRE-REQ")


def _is_column_header(row: tuple) -> bool:
    """Return True if this is the column header row."""
    return row[COL_SCRIPT_ID] == "Script ID"


def _cell(row: tuple, idx: int) -> str:
    """Safely extract a cell value as a stripped string."""
    val = row[idx] if idx < len(row) else None
    return str(val).strip() if val is not None else ""


def parse_workbook(path: str) -> List[TestScenario]:
    """Parse an EX3 RCM Excel workbook and return all test scenarios.

    Args:
        path: Absolute or relative path to the .xlsx file.

    Returns:
        List of TestScenario objects, each containing their TestStep list.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Find the main test sheet
    sheet_name = next(
        (s for s in wb.sheetnames if "end-to-end" in s.lower() or "test" in s.lower()),
        wb.sheetnames[0]
    )
    ws = wb[sheet_name]

    scenarios: dict[str, TestScenario] = {}
    current_scenario_name: str = ""
    current_role: str = ""
    step_counter: dict[str, int] = {}

    for row in ws.iter_rows(values_only=True):
        # Skip fully empty rows
        if not any(c is not None for c in row):
            continue

        # Skip the title rows and column header row
        if _is_column_header(row) or _is_section_header(row):
            continue

        # Skip rows where Script ID is missing or non-string/non-step
        script_id = _cell(row, COL_SCRIPT_ID)
        if not script_id:
            continue

        # Carry forward scenario name and role from first row of each scenario
        scenario_name = _cell(row, COL_SCENARIO)
        if scenario_name:
            current_scenario_name = scenario_name
        role = _cell(row, COL_ROLE)
        if role:
            current_role = role

        action = _cell(row, COL_ACTION)
        test_data = _cell(row, COL_TEST_DATA)
        expected = _cell(row, COL_EXPECTED)
        notes = _cell(row, COL_COMMENTS)

        # Derive the module from the script ID prefix (e.g. LOGIN, RCM)
        module = script_id.split("-")[0] if "-" in script_id else script_id

        if script_id not in scenarios:
            scenarios[script_id] = TestScenario(
                scenario_id=script_id,
                name=current_scenario_name,
                module=module,
                role=current_role,
                steps=[],
            )
            step_counter[script_id] = 0

        step_counter[script_id] += 1
        step_id = f"{script_id}-{step_counter[script_id]:02d}"

        scenarios[script_id].steps.append(TestStep(
            step_id=step_id,
            action=action,
            target="",       # populated by runner based on action text
            expected_result=expected,
            test_data=test_data,
            notes=notes,
        ))

    wb.close()
    return list(scenarios.values())
